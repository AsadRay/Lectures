import openpyxl
import openpyxl.workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
#Reading Excel Documents 
#opening Excel Documents with openpyxl
wb = openpyxl.load_workbook('/home/asaduzzaman-rayhan/Documents/Example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet3']
print(sheet)
print(type('Sheet'))
print(sheet.title)

print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1'] # Get another cell from the sheet.
print(c.value)
#this is how we can get the row column and value from the cell
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))
print(sheet.cell(row=1, column=2).value)
for i in range (1, 8, 2):
    print(i,sheet.cell(row=i, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

#Converting Between Column Letters and Numbers
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))


print(tuple(sheet['A1':'C3']))
for rowOfCellobjects in sheet['A1':'C3']:
    for cellObj in rowOfCellobjects:
        print(cellObj.coordinate, cellObj.value)
        print('---END OF NOW---')

 
    #Import the openpyxl module.
    #Call the openpyxl.load_workbook() function.
    #Get a Workbook object.
    #Use the active or sheetnames attributes.
    #Get a Worksheet object.
    #Use indexing or the cell() sheet method with row and column keyword arguments.
    #Get a Cell object.
    #Read the Cell objects value attribute.


#Creating and saving Excel Documents
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
wb.save('example_copy.xlsx')#---> save the WorkBook

#creating and removing sheets
print(wb.create_sheet()) #---> add a new sheet
print(wb.sheetnames)

#create a new sheet at index 0
wb.create_sheet(index=0, title='Forst Sheet')
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
del wb['Forst Sheet']
print(wb.sheetnames)

#writing values to cells
wb = openpyxl.workbook()
sheet = wb['Middle Sheet']
sheet['A1'] = 'Hello World!'
print(sheet['A1'].value)      

#Formulas
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # Set the formula.
print(wb.save('writeFormula.xlsx'))

#Setting Row Height and Column Width
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
# Set the height and width:
wb = openpyxl.Workbook()
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

#Merging and umerging cells
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')  # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')  # Merge these two cells.
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')
